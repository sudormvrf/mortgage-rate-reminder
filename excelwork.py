from openpyxl import load_workbook
import datetime
import glob
import os


list_xlsx = glob.glob("xlsx_files/*.xlsx")
latest_xlsx = max(list_xlsx, key=os.path.getctime)
workbook = load_workbook(filename=latest_xlsx)
sheet = workbook['Full History']
row = sheet.max_row - 5
date = sheet[f'A{row}'].value
rate = sheet[f'B{row}'].value

str_date = date.strftime("%m/%d/%Y")
push_message = f"The latest average weekly mortgage rate from Freddie Mac was {rate}%, released on {str_date}."

